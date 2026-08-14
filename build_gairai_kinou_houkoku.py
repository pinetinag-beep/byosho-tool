"""外来機能報告（厚生労働省公表オープンデータ）の取り込み。

病床機能報告の姉妹データで、「オープンデータ医療機関コード」で医療機関を一意に
識別できる（医療機関コードの桁数だけ異なる。下記「医療機関コードの正規化」参照）。
年度ごとに、以下の3種類の「様式」ファイルとして公表されている：

  - 外来様式１（年間値）: 診療科構成。令和7年度以降は常勤/非常勤医師数・
    看護師数・CT/MRI等の設備保有状況も同じファイルに含まれる（病床機能報告の
    病院票に相当する情報が統合された）。
  - 外来様式１（月別値）: 初診/紹介/逆紹介患者数など、月別の紹介動向。
  - 外来様式２: 紹介受診重点外来（医療資源を重点的に活用する外来）の
    実施状況。報告月0＝年間値、1〜12＝各月、という行持ち（病床機能報告のDPC
    データ等と違い、最初からこの形式で1ファイルにまとまっている）。

【令和4年度だけデータ構成が異なる（要注意）】
　令和4年度は上記3種類に分かれておらず、様式１相当が「報告様式１」という
　1ファイルに、部門フラグと月別指標がワイド形式（列名に「令和4年4月」等の
　月がそのまま埋め込まれる。例：「初診患者数 令和4年4月」）で同居している。
　`_reshape_r4_form1()` でこれを他の年度と同じロング形式（年間値・月別値の
　2テーブル）に変換して揃える。様式２相当（「報告様式２」）は他年度と同じ
　報告月0〜12のロング形式のため、特別な変換は不要（列数が少ないだけ）。

　さらに令和4年度の報告様式１は、月別指標のブロックによって参照期間が
　異なる点にも注意（原データの列名から判明。病床機能報告の稼働率が
　「許可病床数は7/1時点・在棟延べ数は前年度実績」という異なる期間を
　組み合わせているのと同種の事情）：
     - 初診/紹介/逆紹介患者数・紹介率・逆紹介率　　　→ 令和4年4月〜令和5年3月
     - 休日/夜間受診・救急搬送等（col167以降）　　　　→ 令和3年4月〜令和4年3月
   本スクリプトは値をそのまま報告月1〜12として格納するのみで、この期間の
   ズレを補正・統一しない（CLAUDE.md方針: 独自解釈による指標の補正はしない）。

【医療機関コードの正規化】
　外来機能報告の「オープンデータ医療機関コード」は10桁（先頭ゼロ付き。例:
　"0101010042"）。一方、病床機能報告データ（data_cache.parquet）の
　「医療機関コード」列は9桁のもの・10桁のものが混在している（先頭が
　ゼロになる医療機関は9桁のまま格納されているため）。従って突合する際は
　両者を `.str.zfill(10)` で10桁ゼロ埋めしてから突き合わせること
　（例: 病床機能報告"101010042".zfill(10) == 外来機能報告"0101010042"）。
　本スクリプトの出力ではオープンデータ医療機関コードは原本のまま10桁で保持する。

出力: gairai_form1_annual.parquet / gairai_form1_monthly.parquet / gairai_form2.parquet
（3テーブルとも「報告年度」列で年度を保持。年度によって存在しない列はNaNになる）

使い方:
    python build_gairai_kinou_houkoku.py 外来機能報告
"""
import argparse
import os
import re

import openpyxl
import pandas as pd

HEADER_ROW = 5
DATA_START_ROW = 7

_YEAR_DIR_RE = re.compile(r"令和(\d+)年")
_MONTH_SUFFIX_RE = re.compile(r"令和\d+年(\d{1,2})月$")

# 令和4年度「報告様式１」の部門フラグ列（識別列の直後、月別指標ブロックの手前）
_DEPT_PREFIX = "外来を行っている診療科"


def _reiwa_dir_to_year(dirname: str) -> int:
    """'令和7年' → 2025 のように、報告年度（西暦）へ変換する。"""
    m = _YEAR_DIR_RE.search(dirname)
    if not m:
        raise ValueError(f"年度フォルダ名から令和年度を判定できません: {dirname}")
    return int(m.group(1)) + 2018


def _dedupe_columns(header_row):
    """同名列が複数回登場する場合（初診/再診で同じ項目名が繰り返される等、
    原データの仕様）に、2回目以降へ __1, __2... を付けて一意化する。
    """
    seen = {}
    cols = []
    for h in header_row:
        h = h.strip() if isinstance(h, str) else h
        if not h:
            h = "_blank"
        if h in seen:
            seen[h] += 1
            cols.append(f"{h}__{seen[h]}")
        else:
            seen[h] = 0
            cols.append(h)
    return cols


def _read_form(path: str, sheet_name: str = None) -> pd.DataFrame:
    """5行目=見出し・6行目=必須区分・7行目以降=データ、という病床機能報告と
    共通の様式構造でExcelを読む（load_mhlw_byosho_extendedと同じ考え方）。
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    header_row = next(sh.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    cols = _dedupe_columns(header_row)

    rows = []
    for row in sh.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row or row[1] in (None, ""):  # オープンデータ医療機関コード列が空の行はスキップ
            continue
        rows.append(row)
    df = pd.DataFrame(rows, columns=cols)
    wb.close()
    return df


def _find_monthly_blocks(columns):
    """「〇〇（年間）」列の直後に12個の「...令和X年Y月」列が続くブロックを検出する。
    戻り値: [(metric_name, [12個の月別列名（4月始まり）]), ...]
    """
    blocks = []
    n = len(columns)
    i = 0
    while i < n:
        col = columns[i]
        if isinstance(col, str) and col.endswith("（年間）"):
            month_cols = []
            j = i + 1
            while j < n and len(month_cols) < 12:
                c = columns[j]
                if isinstance(c, str) and _MONTH_SUFFIX_RE.search(c):
                    month_cols.append(c)
                    j += 1
                else:
                    break
            if len(month_cols) == 12:
                metric = col[: -len("（年間）")]
                blocks.append((metric, month_cols))
                i = j
                continue
        i += 1
    return blocks


def _reshape_r4_form1(df: pd.DataFrame):
    """令和4年度「報告様式１」（ワイド形式）を、他年度と同じ
    (年間値テーブル, 月別値テーブル) のロング形式に変換する。
    """
    id_cols = list(df.columns[:12])  # 病診区分 〜 設置主体
    dept_cols = [c for c in df.columns if isinstance(c, str) and c.startswith(_DEPT_PREFIX)]
    annual_df = df[id_cols + dept_cols].copy()

    blocks = _find_monthly_blocks(list(df.columns))
    monthly_frames = []
    for month_pos in range(12):
        month_row = df[id_cols].copy()
        # 月別指標ブロックの列名（例: "初診患者数 令和4年4月"）が示す実際の暦月を
        # 抽出して報告月とする（4月始まりの並びだが、念のため列名から都度読む）
        _, first_block_months = blocks[0]
        month_num = int(_MONTH_SUFFIX_RE.search(first_block_months[month_pos]).group(1))
        month_row["報告月"] = month_num
        for metric, month_cols in blocks:
            month_row[metric] = df[month_cols[month_pos]].values
        monthly_frames.append(month_row)
    monthly_df = pd.concat(monthly_frames, ignore_index=True)
    return annual_df, monthly_df


def _stringify_object_columns(df: pd.DataFrame) -> pd.DataFrame:
    """オブジェクト型（int/strが混在する）列を文字列に統一する。

    病床機能報告・DPC・様式2と同じく、このデータも年間10件以下の値を
    `*` として報告する（実データで確認済み）。部門フラグ列も"〇"/"-"の
    文字列。数値としての集計・マスク値(*)の解釈は用途が決まってから
    行う方針のため、ここでは型エラーを避けつつ原本の値をそのまま保持する。
    """
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].map(lambda v: str(v) if v is not None else None)
    return df


_LOW_CARDINALITY_COLS = [
    "病診区分", "都道府県コード", "二次医療圏コード", "二次医療圏名",
    "構想区域コード", "構想区域名称", "市区町村コード", "市区町村名称", "設置主体",
]


def _categorize_low_cardinality(df: pd.DataFrame) -> pd.DataFrame:
    """ユニーク数が少ない列（都道府県・二次医療圏等）はcategory型にして
    メモリを抑える（施設基準届出データで確立済みの最適化方針。CLAUDE.md参照）。
    """
    for col in _LOW_CARDINALITY_COLS:
        if col in df.columns:
            df[col] = df[col].astype("category")
    return df


def _load_year(year_dir: str) -> dict:
    """1年度分のフォルダから (annual_df, monthly_df, form2_df) を組み立てる。"""
    year = _reiwa_dir_to_year(os.path.basename(year_dir.rstrip("/")))
    files = sorted(f for f in os.listdir(year_dir) if f.endswith(".xlsx"))

    def _sheets_of(fname):
        wb = openpyxl.load_workbook(os.path.join(year_dir, fname), read_only=True)
        names = wb.sheetnames
        wb.close()
        return names

    # 「別紙_...データ定義.xlsx」（用語集）と、R7フォルダにだけある単体の
    # データ定義ファイル（シート名が定義書と同じ構成）は実データではないため除外
    def _is_definition_file(fname):
        if "データ定義" in fname:
            return True
        sheets = _sheets_of(fname)
        return set(sheets) >= {"外来様式１ (年間値)", "外来様式１ (月別値)", "外来様式２"}

    data_files = [f for f in files if not _is_definition_file(f)]

    annual_df = monthly_df = form2_df = None

    if year == 2022:  # 令和4年度：報告様式１・報告様式２の2ファイル構成
        for f in data_files:
            sheet = _sheets_of(f)[0]
            if sheet == "報告様式１":
                raw = _read_form(os.path.join(year_dir, f))
                annual_df, monthly_df = _reshape_r4_form1(raw)
            elif sheet == "報告様式２":
                form2_df = _read_form(os.path.join(year_dir, f))
    else:
        for f in data_files:
            sheet = _sheets_of(f)[0]
            if "月別" in sheet:
                monthly_df = _read_form(os.path.join(year_dir, f))
            elif "年間" in sheet:
                annual_df = _read_form(os.path.join(year_dir, f))
            else:
                # シート名が「外来様式2」「外来様式２」「Sheet1」等、年度によって
                # 表記が揺れる（全角/半角の「２」・汎用名等）ため、月別・年間の
                # いずれでもないファイルを様式２として扱う
                form2_df = _read_form(os.path.join(year_dir, f))

    for df in (annual_df, monthly_df, form2_df):
        if df is not None:
            df.insert(0, "報告年度", year)

    missing = [name for name, df in
               [("annual", annual_df), ("monthly", monthly_df), ("form2", form2_df)] if df is None]
    if missing:
        raise RuntimeError(f"{year_dir}: 次のテーブルが見つかりませんでした: {missing}")

    return {"annual": annual_df, "monthly": monthly_df, "form2": form2_df}


def build(root_dir: str):
    year_dirs = sorted(
        os.path.join(root_dir, d) for d in os.listdir(root_dir)
        if os.path.isdir(os.path.join(root_dir, d)) and _YEAR_DIR_RE.search(d)
    )
    if not year_dirs:
        raise RuntimeError(f"{root_dir} 直下に「令和X年」フォルダが見つかりません")

    annuals, monthlies, form2s = [], [], []
    for yd in year_dirs:
        print(f"読み込み中: {yd}")
        parts = _load_year(yd)
        annuals.append(parts["annual"])
        monthlies.append(parts["monthly"])
        form2s.append(parts["form2"])
        print(f"  年間値: {len(parts['annual'])}行 / 月別値: {len(parts['monthly'])}行 / 様式2: {len(parts['form2'])}行")

    annual_all = _categorize_low_cardinality(_stringify_object_columns(pd.concat(annuals, ignore_index=True)))
    monthly_all = _categorize_low_cardinality(_stringify_object_columns(pd.concat(monthlies, ignore_index=True)))
    form2_all = _categorize_low_cardinality(_stringify_object_columns(pd.concat(form2s, ignore_index=True)))

    annual_all.to_parquet("gairai_form1_annual.parquet", index=False)
    monthly_all.to_parquet("gairai_form1_monthly.parquet", index=False)
    form2_all.to_parquet("gairai_form2.parquet", index=False)

    print("\n完了:")
    print(f"  gairai_form1_annual.parquet : {len(annual_all)}行 {len(annual_all.columns)}列")
    print(f"  gairai_form1_monthly.parquet: {len(monthly_all)}行 {len(monthly_all.columns)}列")
    print(f"  gairai_form2.parquet        : {len(form2_all)}行 {len(form2_all.columns)}列")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="外来機能報告オープンデータの取り込み")
    parser.add_argument("dir", help="令和X年フォルダを含む親ディレクトリ（例: 外来機能報告）")
    args = parser.parse_args()
    build(args.dir)
