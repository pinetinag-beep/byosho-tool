#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
byosho_file_R4〜R7 フォルダに紛れている様式2（手術データ・7〜8地域ファイル）を
読み込んで surgery_cache.parquet に取り込む。

【重要・2026年8月にラベル方針を変更】様式2ファイルの実績期間（診療期間）は、
フォルダ名（＝病棟票・病院票の報告年度）より1年古い（病床機能報告は7月1日
時点のward census と、その時点までに確定している直近1年間＝前年度の手術
実績をセットで提出する仕組みのため）。実例: byosho_file_R7 内の様式2ファイル
は「令和6年4月から令和7年3月診療分」＝令和6年度の実績。

以前はこのファイル内埋め込みテキストから実際の診療期間を自動判定し、それを
「報告年度」ラベルとして使っていた（byosho_file_R7 → 2024とラベル）。しかし
これだと同じ「2025年度」を選んで病床機能報告と様式2を見比べたときに様式2側
だけ「まだ公開されていない」ように見えてしまい紛らわしいとユーザーから指摘
された。**現在はフォルダ名が示す病床機能報告自体の報告年度をそのままラベル
として使う**方針に変更した（byosho_file_R7 → 2025とラベル。実績期間が実際
には1年古いことはラベルには反映しない）。ファイル内埋め込みテキストの実際の
診療期間は診断用に表示するだけで、ラベルには使わない。

使い方:
    python build_yoshiki2_from_byosho.py byosho_file_R7 \
        --exclude 001717798 001717800 001717801 001717802 001717803 001717804 001717805 001717806
"""
import argparse
import glob
import re
from pathlib import Path

import pandas as pd
import openpyxl

from data_processor import load_mhlw_yoshiki2

BASE = Path(__file__).parent
SURGERY_PARQUET = BASE / "surgery_cache.parquet"

_PERIOD_RE = re.compile(r"令和(\d+)年(\d+)月から令和(\d+)年(\d+)月診療分")
_FOLDER_RE = re.compile(r"R(\d+)")


def _folder_report_year(folder_name: str) -> int:
    """フォルダ名（例: byosho_file_R7）から病床機能報告の報告年度（西暦）を
    導出する。R7 → 令和7年度 → 2025。"""
    m = _FOLDER_RE.search(folder_name)
    if not m:
        raise ValueError(f"フォルダ名から報告年度を判定できませんでした: {folder_name}")
    return int(m.group(1)) + 2018  # 令和1年 = 2019


def _detect_period_year(path: str) -> int | None:
    """ファイル先頭の埋め込みテキストから実際の診療期間の開始年度（西暦）を
    判定する（診断・整合性チェック用。ラベルには使わない）。"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for v in row:
            if not v:
                continue
            m = _PERIOD_RE.search(str(v))
            if m:
                reiwa_year = int(m.group(1))
                return reiwa_year + 2018
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("folder", help="byosho_file_R4 等のフォルダ名")
    ap.add_argument("--exclude", nargs="+", required=True,
                     help="様式1（病棟票・病院票）で使用済みのファイルIDプレフィックス一覧")
    args = ap.parse_args()

    folder = BASE / args.folder
    year = _folder_report_year(args.folder)
    all_files = sorted(glob.glob(str(folder / "*.xlsx")))
    targets = [f for f in all_files if not any(Path(f).stem.startswith(e) for e in args.exclude)]
    if not targets:
        print("対象ファイルが見つかりません（--excludeの指定を確認してください）")
        return

    print(f"対象ファイル数: {len(targets)}　報告年度（フォルダ由来）: {year}")
    parts = []
    for f in targets:
        period_year = _detect_period_year(f)
        if period_year is not None and period_year != year:
            print(f"  {Path(f).name}: 実際の診療期間は{period_year}年度（ラベルは{year}年度のまま使用）")
        with open(f, "rb") as fh:
            df = load_mhlw_yoshiki2(fh.read(), year=year)
        print(f"  {Path(f).name}: {len(df)}行")
        parts.append(df)

    combined = pd.concat(parts, ignore_index=True)
    before = len(combined)
    combined = combined.drop_duplicates(subset=["医療機関名", "都道府県名"])
    print(f"\n報告年度 {year}: {before}行 → 重複除去後 {len(combined)}行")

    if SURGERY_PARQUET.exists():
        existing = pd.read_parquet(SURGERY_PARQUET)
        if year in existing["報告年度"].unique():
            print(f"既存の{year}年度データ（{(existing['報告年度']==year).sum()}行）を置き換えます")
            existing = existing[existing["報告年度"] != year]
        merged = pd.concat([existing, combined], ignore_index=True)
    else:
        merged = combined

    merged.to_parquet(SURGERY_PARQUET, index=False)
    print(f"\n=== 完了 === {SURGERY_PARQUET} 合計 {len(merged)}行")
    print(merged["報告年度"].value_counts().sort_index())


if __name__ == "__main__":
    main()
